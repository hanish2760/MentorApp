from django.contrib.auth import authenticate
import django.contrib.auth.mixins
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy
from django.views  import View
from django.shortcuts import render, get_object_or_404,redirect
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import*
from rest_framework.parsers import JSONParser
from rest_framework.views import APIView

from onlineapp.models import*
from django.forms import *
from onlineapp.forms.Auth import *
from onlineapp.forms.Login import *
from django.contrib.auth import *


# super user hanish2760 lostlost

class Home(View):
    def get(self,request):
        return render(request,template_name='home.html')

class CollegeView(django.contrib.auth.mixins.LoginRequiredMixin, View):
    login_url = '/login/'
    # we can import packages anywhere lol
    def get_context_data(self,**kwargs):
        context = super(CollegeView, self).get_comtext_data(**kwargs)
        context.update({
            'user_permissions': self.request.user.get_all_permissions(),
        })
    def get(self,request,*args,**kwargs):

        import ipdb
     #debugging point!   ipdb.set_trace()
        collegelist=College.objects.all()
        # ipdb.set_trace()
        return render(
            request,
            template_name='collegeList.html',
            context={'col':collegelist},
        )
        # using short cut render function whicch does same worl as in views.backuo

#ListView is a short cut where the listview is going to understand that we are calling the entire list like  College.objects.all().

class  CollegeListView(django.contrib.auth.mixins.LoginRequiredMixin, ListView):
    login_url = '/login/'
    model=College

    context_object_name = 'col'
    template_name = 'collegeList.html'

#pagination happens in listvviewc

    def get_context_data(self,**kwargs):

        context=super(CollegeListView,self).get_context_data(**kwargs)
        context.update({
            'user_permissions': self.request.user.get_all_permissions(),
        })
        return context


#https://docs.djangoproject.com/en/2.0/ref/class-based-views/generic-display/#django.views.generic.list.BaseListView
class CollegeDetailsView(django.contrib.auth.mixins.LoginRequiredMixin, DetailView):
    login_url = '/login/'
   
    model=College
    template_name = 'collge_Details.html'

    def get_object(self, queryset=None):

        return get_object_or_404(College,**self.kwargs) #**{'pk':self.kwargs.get('id')}

    def get_context_data(self, **kwargs):

        context=super(CollegeDetailsView,self).get_context_data(**kwargs)
        context.update({
            'user_permissions': self.request.user.get_all_permissions(),
        })
        college=context.get('object')
#  amde a change here id is added.
        std=self.model.objects.filter(id=college.id).values('student__mocktest1__total', 'student__name', 'acronym','id')
        students=self.model.objects.filter(id=college.id).values('student__mocktest1__total','student__name','acronym','student__id','student__college__id')
        students = list(
            college.student_set.values('id', 'name', 'email', 'mocktest1__total').order_by('-mocktest1__total'))

        #.order_by('-student____mocktest1__total')
        #self.model.objects."query"
       # students=list(college.student.order_by('-mocktest1__total'))
        context.update(
            {
                'students':students
            }
        )

        return context

class AddCollege(ModelForm):
    class Meta:
        model = College
        exclude = ['id']
        widgets = {
            'name': TextInput(attrs={'class':'form-control','placeholder':'Enter College name'}),
            'location': TextInput(attrs={'class': 'form-control', 'placeholder': 'Enter College Location'}),
            'acronym': TextInput(attrs={'class': 'form-control', 'placeholder': 'Enter Acronym'}),
            'contact': TextInput(attrs={'class': 'form-control', 'placeholder': 'Enter Contact Email'}),
        }

class CreateCollegeView(django.contrib.auth.mixins.LoginRequiredMixin, CreateView):
    login_url = '/login/'
    model = College
    form_class = AddCollege
    template_name = "college_form.html"
    success_url = reverse_lazy('collegesList')
    def get_context_data(self, **kwargs):
        context = super(CreateCollegeView, self).get_context_data(**kwargs)
        context.update({
            'user_permissions': self.request.user.get_all_permissions(),
        })
        return context

class AddStudent(ModelForm):
    class Meta:
        model = Student
        exclude = ['id','college','dob']
        widgets = {
            'name': TextInput(attrs={'class':'form-control','placeholder':'Enter Student name'}),
            #'dob':DateInput(attrs={'class':'form_control','placeholder':'Puttina rooju cheppumu'}),
            'email':EmailInput(attrs={'class':'form_control','placeholder':'Email id-pls jara correct pls'}),
            'dropped_out':CheckboxInput(attrs={'class':'form-control','placeholder':'is dropped out?'}),

            'db_folder':TextInput(attrs={'class':'form_control','placeholder':'Enter Db_folder'}),
        }


class AddMocktest(ModelForm):
    class Meta:
        model = MockTest1
        exclude = ['id', 'students', 'total']
        widgets = {
            'problem1': NumberInput(attrs={'class': 'form-control', 'placeholder': 'Enter marks 1'}),
            'problem2': NumberInput(attrs={'class': 'form-control', 'placeholder': 'Enter marks 2'}),
            'problem3': NumberInput(attrs={'class': 'form-control', 'placeholder': 'Enter marks 3'}),
            'problem4': NumberInput(attrs={'class': 'form-control', 'placeholder': 'Enter marks 4'}),
        }



class CreateStudentView(django.contrib.auth.mixins.LoginRequiredMixin, CreateView):
    login_url = '/login/'
    model=Student
    form_class = AddStudent
    template_name = "student_form.html"


    def get_context_data(self, **kwargs):
        context=super(CreateStudentView,self).get_context_data(**kwargs)
        student_form=context.get('form')
        marks_form=AddMocktest()
        context.update({
            'user_permissions': self.request.user.get_all_permissions(),
        })
        context.update(
            {
                'student': student_form,
                'mock': marks_form,
            }
        )
        return context
    def post(self, request, *args, **kwargs):
        college = get_object_or_404(College, pk=kwargs.get('pk'))

        student_form = AddStudent(request.POST)
        mocktest_form = AddMocktest(request.POST)

        if student_form.is_valid():
            student = student_form.save(commit=False)
            student.college = college
            student.save()

            if mocktest_form.is_valid():
                mocktest = mocktest_form.save(commit=False)
                mocktest.total = sum(mocktest_form.cleaned_data.values())
                mocktest.students = student
                mocktest.save()
        return redirect('marks', college.id)


class EditCollegeView(django.contrib.auth.mixins.LoginRequiredMixin, UpdateView):
    login_url = '/login/'
    model = College
    fields = [
        'name', 'acronym', 'location', 'contact'
    ]
    template_name = 'college_form.html'
    success_url = reverse_lazy('collegesList')
    def get_context_data(self, **kwargs):
        context = super(EditCollegeView, self).get_context_data(**kwargs)
        context.update({
            'user_permissions': self.request.user.get_all_permissions(),
        })
        return context


class DeleteCollegeView(django.contrib.auth.mixins.LoginRequiredMixin, DeleteView):
    login_url = '/login/'
    model = College
    template_name = "college_confirm_delete.html"
    success_url = reverse_lazy('collegesList')
    def get_context_data(self, **kwargs):
        context = super(DeleteCollegeView, self).get_context_data(**kwargs)
        context.update({
            'user_permissions': self.request.user.get_all_permissions(),
        })
        return context

class EditStudentView(django.contrib.auth.mixins.LoginRequiredMixin, UpdateView):
    login_url = '/login/'
    model = Student
    template_name = "student_form.html"
    form_class = AddStudent

    def get_context_data(self, **kwargs):
        context = super(EditStudentView, self).get_context_data(**kwargs)
        student_form = context.get('form')
        student = context['object']
        #mocktest_form = AddMocktest(instance=get_object_or_404(MockTest1,**kwargs)) when student dnt writee exam
        try:
            mock=MockTest1.objects.get(student=student)
        except:
            mock=None
        mocktest_form=AddMocktest(instance=mock)

        context.update({
            'user_permissions': self.request.user.get_all_permissions(),
        })
        context.update({
            'student': student_form,
            'mock': mocktest_form,
        })
        return context

    def post(self, request, *args, **kwargs):
        #ipdb.set_trace()

        acr = kwargs['acronym']
        college_object = College.objects.get(acronym=acr)
        mocktest = get_object_or_404(MockTest1, student_id=self.kwargs.get('pk'))
        student = Student.objects.get(id=self.kwargs.get('pk'))
        student_form = AddStudent(request.POST, instance=student)
        mock_form = AddMocktest(request.POST, instance=mocktest)
        if student_form.is_valid():
            student.save()
            if mock_form.is_valid():
                mocktest.totals = sum(mock_form.cleaned_data.values())
                mocktest.save()
        return redirect("college_details", college_object.acronym)

class DeleteStudentView(django.contrib.auth.mixins.LoginRequiredMixin, DeleteView):
    login_url = '/login/'
    model = Student
    template_name = "college_confirm_delete.html"
    def get_context_data(self, **kwargs):
        context = super(DeleteStudentView, self).get_context_data(**kwargs)
        context.update({
            'user_permissions': self.request.user.get_all_permissions(),
        })
        return context

    def post(self, request, *args, **kwargs):

        std=self.model.objects.get(id=kwargs['pk'])

        collegeid=std.college.id
        std.delete()

        return redirect('marks',collegeid)





""""
check it out 
redirect
receiver something  


"""


# write this is in new view function

class  singnUpView(View):
    def get(self,request):
        signform=signupform()
        # refer documnet
        return render(
            request,
            template_name="signup_form.html",
            context={
                'form'  : signform,
            }
        )
    def post(self,request):

        form=signupform(request.POST)
        if form.is_valid() :

            if authenticate(request,username=form.cleaned_data['username'],
                            password=form.cleaned_data['password'], ):
                return HttpResponse("User Already Exits!")

            User.objects.create_user(username=form.cleaned_data['username'],password=form.cleaned_data['password'])
            user =authenticate(
                request,
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password'],
            )
            if(user is not None):
                user.save()
                # login(request,user)
                return redirect('collegesList')

            # else:
            #     return redirect('collegesList')


class LoginView(View):
    def get(self,request):
        loginform=loginForm()
        # import ipdb
        # ipdb.set_trace()

        return render(
            request,
            template_name="login.html",
            context={
            'form':loginform,
            },
        )

    def post(self, request):
        login_form = loginForm(request.POST)
        if login_form.is_valid():
            user = authenticate(
                request,
                username=login_form.cleaned_data['username'],
                password=login_form.cleaned_data['password'],
            )
            if user:
                login(request, user)
                return redirect('collegesList')
            else:
                return HttpResponse("Login Failed !")


def logout_user(request):
    logout(request)
    return redirect('login')

import openpyxl
from openpyxl import load_workbook

def read_from_xl2(workbookName, sheetName):
    '''this function reads data from a specified sheet and returns a 2d list of data'''
    wb =  load_workbook(workbookName)
    sheetsrc = wb[sheetName]

    rows = sheetsrc.max_row
    cols = sheetsrc.max_column

    data = []
    for r in range(0, rows):
        data.append([])

    for r in range(rows):
        for c in range(cols):
            cell = sheetsrc.cell(row=r + 1, column=c + 1)
            data[r].append(cell.value)

    return data
def populatedb2():
    marks = read_from_xl2("result1.xlsx", "TableData")[1:]
    colleges = read_from_xl2("students.xlsx", "Colleges")[1:]
    current = read_from_xl2("students.xlsx", "Current")[1:]
    deletions = read_from_xl2("students.xlsx", "Deletions")[1:]

    collegeDict = dict()
    stdFolderDict = dict()

    for college in colleges:
        c = College(name=college[0], location=college[2], acronym=college[1], contact=college[3])
        c.save()
        collegeDict[college[1]] = c

    for student in current:
        try:
            collegeDict[student[1]]
        except Exception:
            print(student[1], "not found")
        else:
            s = Student(name=student[0], dob=None, email=student[2], db_folder=student[3], dropped_out=False,
                        college=collegeDict[student[1]])
            s.save()
            stdFolderDict[str(student[3]).lower()] = s

    for dStudent in deletions:
        s = Student(name=dStudent[0], dob=None, email=dStudent[2], db_folder=dStudent[3], dropped_out=True,
                    college=collegeDict[dStudent[1]])
        s.save()

    for std in marks:
        dbname = str(str(std[0]).split('_')[2]).lower()
        try:
            stdFolderDict[dbname]
        except Exception:
            print(dbname, " not found")
        else:
            m = MockTest1(problem1=str(std[1]), problem2=str(std[2]), problem3=str(std[3]), problem4=str(std[4]),
                          total=str(std[5]), students=stdFolderDict[dbname])
            m.save()


@csrf_exempt
def upload_data(request):
    populatedb2()
    return HttpResponse("hello")
