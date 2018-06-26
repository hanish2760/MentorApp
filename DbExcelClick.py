import os
import django
import requests
import MySQLdb
from bs4 import BeautifulSoup
import openpyxl as xl
import click



os.environ.setdefault("DJANGO_SETTINGS_MODULE", "onlineClass.settings")
django.setup()

from onlineapp.models import *


#####################################################################################
def extractWebData(soup):
    tableHeads = [x.text.strip("</th>").strip("<th>") for x in soup.find_all("th")][1:]
    tableData = [x.text.strip("</td>").strip("<td>").lstrip('0123456789') for x in soup.find_all("tr")][1:]


    return (tableHeads,tableData)

def storeWebDataToExcel(destName, tableHeads, tableData):
    workbook = xl.Workbook()
    workbook.create_sheet("TableData")
    workbook.save(destName)

    sheet = workbook['TableData']

    for col in range(6):
        cell = sheet.cell(row=1, column=col + 1)
        cell.value = tableHeads[col]


    for row in range(1,len(tableData)):
        inputs = tableData[row].split()
        for col in range(len(inputs)):
            cell = sheet.cell(row = row+1,column = col+1)
            cell.value = inputs[col]


    workbook.save(destName)

def read_from_xl(workbookName, sheetName):
    '''this function reads data from a specified sheet and returns a 2d list of data'''
    wb = xl.load_workbook(workbookName)
    sheetsrc = wb[sheetName]

    rows = sheetsrc.max_row
    cols = sheetsrc.max_column

    data = []
    for r in range(0, rows):
        data.append([])

    for r in range(rows):
        for c in range(cols):
            cell = sheetsrc.cell(row=r + 1, column=c + 1)
            data[r].append(cell.value)

    return data
#####################################################################################

# print(read_from_xl("students.xlsx","Colleges"))

@click.group()
def cli():
    pass


@cli.command('createdb')
def createDB():
    # a function to create a database and add 2 tables into it
    db = MySQLdb.connect(host="localhost", user="root", passwd="27608678")
    c = db.cursor()
    c.execute("create database tutdb")
    db.commit()
    print('created db')


@cli.command('deletedb')
def deleteDB():
    db = MySQLdb.connect(host="localhost", user="root", passwd="password")
    c = db.cursor()
    c.execute("drop database tutdb")
    db.commit()
    print('deleted db')


@cli.command('populatedb')
@click.argument('stdFile')
@click.argument('mrksFile')
def populatedb(stdfile,mrksfile):
    # request = requests.get(mrksfile).text
    # soup = BeautifulSoup(request, 'html5lib')
    #
    # webData = extractWebData(soup)
    # storeWebDataToExcel("result1.xlsx", webData[0], webData[1])

    marks = read_from_xl("result1.xlsx", "TableData")[1:]
    colleges = read_from_xl(stdfile, "Colleges")[1:]
    current = read_from_xl(stdfile, "Current")[1:]
    deletions = read_from_xl(stdfile, "Deletions")[1:]

    collegeDict = dict()
    stdFolderDict = dict()

    for college in colleges:
        c = College(name=college[0], location=college[2], acronym=college[1], contact=college[3])
        c.save()
        collegeDict[college[1]] = c

    for student in current:
        try:
            collegeDict[student[1]]
        except Exception:
            print(student[1], "not found")
        else:
            s = Student(name=student[0], dob=None, email=student[2], db_folder=student[3], dropped_out=False,
                        college=collegeDict[student[1]])
            s.save()
            stdFolderDict[str(student[3]).lower()] = s

    for dStudent in deletions:
        s = Student(name=dStudent[0], dob=None, email=dStudent[2], db_folder=dStudent[3], dropped_out=True,
                    college=collegeDict[dStudent[1]])
        s.save()


    for std in marks:
        dbname = str(str(std[0]).split('_')[2]).lower()
        try:
            stdFolderDict[dbname]
        except Exception:
            print(dbname, " not found")
        else:
            m = MockTest1(problem1=str(std[1]), problem2=str(std[2]), problem3=str(std[3]), problem4=str(std[4]),
                          total=str(std[5]), students=stdFolderDict[dbname])
            m.save()


@cli.command('cleardata')
def clearData():
    College.delete()
    print('cleardata')


# if __name__ == '__main__':
#    cli()


#c = College(name = "Vasavi college",location = "Hyderabad",acronym = "vce",contact = "vce@gmail.com")

#c.save()
#print('hk')
# createDB()
# populatedb()
#most effective cause the dtabase somewhere stores the count on the root
# c=College.objects.filter(location="Hyderabad").count()
# print(c.query,"  ")

#c=College.objects.order_by('-acronym') decending
# a=College.objects.values('acronym')
# a.sorted(reverse=True)
# a max heap is made and top 5  are selectted
#c=College.objects.order_by('location')[:5]
#SELECT DISTINCT `onlineapp_college`.`location` FROM `onlineapp_college`    [{'location': 'Hyderabad'}, {'location': 'Vizag'}, {'location': 'Bhimavaram'}, {'location': 'Bapatla'}]

""""c=College.objects.values('location').distinct()
# aggregate vs annotate.
print(c.query,"  ",list(c))
for a in c:
    print(a['location'],"   ",College.objects.filter(location=a['location']).count())

not in one query!!

Group by query!
"""""
# from django.db.models import Count
# pubs = College.objects.annotate(num_books=Count('location'))
# print(pubs.query,"  ",pubs)
"""
from django.db.models import Avg, Max, Min
q=MockTest1.objects.aggregate(Min('problem1'))
print(q)
aggregate is not a query set.

orderby vs groupby

annotate : SELECT `onlineapp_college`.`location`, COUNT(`onlineapp_college`.`location`) AS `location__count` FROM `onlineapp_college` GROUP BY `onlineapp_college`.`location` ORDER BY NULL
"""

from django.db.models import Count

#q=College.objects.values('location').annotate(Count('location')) group by col 1(location) then performs a funtion on  th values of the grp

#q=College.objects.annotate(Count('location')) grps by id.
# q=College.objects.values('location').annotate(cnt=Count('location')).order_by('-cnt')
# print(q.query,"  ",q)
#  __ goes to the back link object. student.objects.filter(drop_out=True).values('college__name').distinct()
# #inner joins!
# q=Student.objects.filter(dropped_out=True).values('college__name')
# print(q.query,"  ",q)
# q=Student.objects.filter(college__acronym="vce").values('name')
# print(q.query,"  ",q)
# q=Student.objects.values('college__name','college__acronym').annotate(cnt=Count('name'))
# print(q.query,"  ",q)
#annotate means we are grouping by college name in the above query.
# bottom up appoach above.
# top down.\|/
#q=College.objects.filter(student__dropout=True)

"""different type of joins which is more performant??"""

# q=College.objects.values('acronym').annotate(cnt=Count('student__id'))
# print(q.query,"  ",q)


q=Student.objects.values('name','college__name')

#q=Student.objects.all().annotate()
from django.db.models import Q
#lookup fields....cnt >=10
q=Student.objects.values('college__name').annotate(cnt=Count('id')).order_by('-cnt').filter(cnt__gte=10)
# checckout the query!! student name-college name- total
q=MockTest1.objects.values('total','students__name','students__college__acronym')

q=MockTest1.objects.values('total','students__name','students__college__acronym').filter(total__gte=30).order_by('total')
# avg mac min
# from django.db.models import*
# q=MockTest1.objects.values('students__college__name').annotate(Max('total'),Avg('total'),Min('total'))
#
# q=Student.objects.values('college__id').annotate(Avg('mocktest1__total'))
#  what is happening here ? ? in the for looop

# manager=College.objects
# querysets=College.objects.all()
#
# for s in q:
#     print(s)
#col=

from rest_framework import serializers
#django.setup()
class CollegeSerializer(serializers.Serializer):
    name = serializers.CharField(max_length=128)  # collegename
    location = serializers.CharField(max_length=64)
    acronym = serializers.CharField(max_length=8)
    contact = serializers.EmailField()

col=College.objects.filter(location="Hyderabad")
serializer = CollegeSerializer(col[0])
for c in col:
    serializer=CollegeSerializer(c)
    print(serializer.data)

import threading
import time
